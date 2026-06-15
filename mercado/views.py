from decimal import Decimal
from difflib import SequenceMatcher
from django.contrib import messages
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from openpyxl import Workbook, load_workbook
from .models import (
    SiteMonitorado,
    Categoria,
    Marca,
    FabricanteImportador,
    ProdutoReferencia,
    ProdutoColetado,
    ColetaProduto,
)


def dashboard(request):
    contexto = {
        "total_sites": SiteMonitorado.objects.count(),
        "total_categorias": Categoria.objects.count(),
        "total_marcas": Marca.objects.count(),
        "total_fabricantes": FabricanteImportador.objects.count(),
        "total_referencias": ProdutoReferencia.objects.count(),
        "total_produtos": ProdutoColetado.objects.count(),
        "total_coletas": ColetaProduto.objects.count(),
        "ultimas_coletas": ColetaProduto.objects.select_related(
            "produto",
            "produto__site",
            "produto__marca",
            "produto__categoria",
            "produto__produto_referencia",
        ).order_by("-data_coleta")[:10],
    }

    return render(request, "mercado/dashboard.html", contexto)


def lista_produtos(request):
    produtos = ProdutoColetado.objects.select_related(
        "site",
        "marca",
        "categoria",
        "produto_referencia",
    ).all()

    busca = request.GET.get("busca", "")
    site_id = request.GET.get("site", "")
    marca_id = request.GET.get("marca", "")
    categoria_id = request.GET.get("categoria", "")

    if busca:
        produtos = produtos.filter(
            nome_original__icontains=busca
        ) | produtos.filter(
            codigo_fabricante__icontains=busca
        ) | produtos.filter(
            ean__icontains=busca
        ) | produtos.filter(
            produto_referencia__nome_referencia__icontains=busca
        ) | produtos.filter(
            produto_referencia__codigo_fabricante__icontains=busca
        ) | produtos.filter(
            produto_referencia__ean__icontains=busca
        )

    if site_id:
        produtos = produtos.filter(site_id=site_id)

    if marca_id:
        produtos = produtos.filter(marca_id=marca_id)

    if categoria_id:
        produtos = produtos.filter(categoria_id=categoria_id)

    contexto = {
        "produtos": produtos.distinct(),
        "sites": SiteMonitorado.objects.all(),
        "marcas": Marca.objects.all(),
        "categorias": Categoria.objects.all(),
        "busca": busca,
        "site_id": site_id,
        "marca_id": marca_id,
        "categoria_id": categoria_id,
    }

    return render(request, "mercado/lista_produtos.html", contexto)


def lista_referencias(request):
    referencias = ProdutoReferencia.objects.select_related(
        "marca",
        "categoria",
        "fabricante_importador",
    ).all()

    busca = request.GET.get("busca", "")
    marca_id = request.GET.get("marca", "")
    categoria_id = request.GET.get("categoria", "")
    fabricante_id = request.GET.get("fabricante", "")
    status_validacao = request.GET.get("status_validacao", "")

    if busca:
        referencias = referencias.filter(
            nome_referencia__icontains=busca
        ) | referencias.filter(
            codigo_fabricante__icontains=busca
        ) | referencias.filter(
            ean__icontains=busca
        ) | referencias.filter(
            url_oficial__icontains=busca
        )

    if marca_id:
        referencias = referencias.filter(marca_id=marca_id)

    if categoria_id:
        referencias = referencias.filter(categoria_id=categoria_id)

    if fabricante_id:
        referencias = referencias.filter(fabricante_importador_id=fabricante_id)

    if status_validacao:
        referencias = referencias.filter(status_validacao=status_validacao)

    contexto = {
        "referencias": referencias.distinct(),
        "marcas": Marca.objects.all(),
        "categorias": Categoria.objects.all(),
        "fabricantes": FabricanteImportador.objects.all(),
        "status_choices": ProdutoReferencia.STATUS_VALIDACAO_CHOICES,
        "busca": busca,
        "marca_id": marca_id,
        "categoria_id": categoria_id,
        "fabricante_id": fabricante_id,
        "status_validacao": status_validacao,
    }

    return render(request, "mercado/lista_referencias.html", contexto)

def calcular_diferenca_prazo_vista(preco_atual, preco_prazo):
    if not preco_atual or not preco_prazo:
        return None

    try:
        preco_atual = Decimal(preco_atual)
        preco_prazo = Decimal(preco_prazo)

        if preco_atual <= 0:
            return None

        diferenca = ((preco_prazo - preco_atual) / preco_atual) * Decimal("100")
        return round(diferenca, 2)

    except Exception:
        return None


def montar_texto_parcelamento(coleta):
    if not coleta:
        return "-"

    if coleta.quantidade_parcelas and coleta.valor_parcela:
        return f"{coleta.quantidade_parcelas}x de R$ {coleta.valor_parcela}"

    return "-"

def ranking_mercado(request):
    produtos = ProdutoColetado.objects.select_related(
        "site",
        "marca",
        "categoria",
        "produto_referencia",
    ).filter(
        ativo=True
    )

    busca = request.GET.get("busca", "")
    site_id = request.GET.get("site", "")
    marca_id = request.GET.get("marca", "")
    categoria_id = request.GET.get("categoria", "")
    somente_disponiveis = request.GET.get("somente_disponiveis", "")

    if busca:
        produtos = produtos.filter(
            Q(nome_original__icontains=busca)
            | Q(codigo_site__icontains=busca)
            | Q(codigo_fabricante__icontains=busca)
            | Q(ean__icontains=busca)
            | Q(url__icontains=busca)
        )

    if site_id:
        produtos = produtos.filter(site_id=site_id)

    if marca_id:
        produtos = produtos.filter(marca_id=marca_id)

    if categoria_id:
        produtos = produtos.filter(categoria_id=categoria_id)

    ranking = []

    for produto in produtos:
        ultima_coleta = produto.coletas.order_by("-data_coleta").first()

        if not ultima_coleta:
            continue

        if somente_disponiveis and not ultima_coleta.disponivel:
            continue

        diferenca_prazo_vista = calcular_diferenca_prazo_vista(
            ultima_coleta.preco_atual,
            ultima_coleta.preco_prazo,
        )

        ranking.append({
            "produto": produto,
            "coleta": ultima_coleta,
            "preco_atual": ultima_coleta.preco_atual,
            "preco_antigo": ultima_coleta.preco_antigo,
            "preco_prazo": ultima_coleta.preco_prazo,
            "quantidade_parcelas": ultima_coleta.quantidade_parcelas,
            "valor_parcela": ultima_coleta.valor_parcela,
            "parcelamento": montar_texto_parcelamento(ultima_coleta),
            "diferenca_prazo_vista": diferenca_prazo_vista,
            "ranking_geral": ultima_coleta.ranking_geral,
            "ranking_categoria": ultima_coleta.ranking_categoria,
            "nota_media": ultima_coleta.nota_media,
            "quantidade_avaliacoes": ultima_coleta.quantidade_avaliacoes,
            "disponivel": ultima_coleta.disponivel,
            "data_coleta": ultima_coleta.data_coleta,
        })

    ranking.sort(
        key=lambda item: (
            item["ranking_geral"] if item["ranking_geral"] is not None else 999999,
            item["preco_atual"] if item["preco_atual"] is not None else Decimal("999999999"),
        )
    )

    contexto = {
        "ranking": ranking,
        "total_produtos": len(ranking),
        "sites": SiteMonitorado.objects.all(),
        "marcas": Marca.objects.all(),
        "categorias": Categoria.objects.all(),
        "busca": busca,
        "site_id": site_id,
        "marca_id": marca_id,
        "categoria_id": categoria_id,
        "somente_disponiveis": somente_disponiveis,
    }

    return render(request, "mercado/ranking_mercado.html", contexto)

def precos_referencias(request):
    coletas = ColetaProduto.objects.select_related(
        "produto",
        "produto__site",
        "produto__marca",
        "produto__categoria",
        "produto__produto_referencia",
        "produto__produto_referencia__marca",
        "produto__produto_referencia__categoria",
        "produto__produto_referencia__fabricante_importador",
    ).filter(
        produto__produto_referencia__isnull=False,
        preco_atual__isnull=False,
    )

    busca = request.GET.get("busca", "")
    site_id = request.GET.get("site", "")
    marca_id = request.GET.get("marca", "")
    categoria_id = request.GET.get("categoria", "")
    data_inicio = request.GET.get("data_inicio", "")
    data_fim = request.GET.get("data_fim", "")

    if busca:
        coletas = coletas.filter(
            Q(produto__produto_referencia__nome_referencia__icontains=busca)
            | Q(produto__produto_referencia__codigo_fabricante__icontains=busca)
            | Q(produto__produto_referencia__ean__icontains=busca)
            | Q(produto__nome_original__icontains=busca)
            | Q(produto__codigo_fabricante__icontains=busca)
            | Q(produto__ean__icontains=busca)
        )

    if site_id:
        coletas = coletas.filter(produto__site_id=site_id)

    if marca_id:
        coletas = coletas.filter(produto__produto_referencia__marca_id=marca_id)

    if categoria_id:
        coletas = coletas.filter(produto__produto_referencia__categoria_id=categoria_id)

    if data_inicio:
        coletas = coletas.filter(data_coleta__date__gte=data_inicio)

    if data_fim:
        coletas = coletas.filter(data_coleta__date__lte=data_fim)

    grupos = {}

    for coleta in coletas.order_by("produto__produto_referencia__nome_referencia", "data_coleta"):
        referencia = coleta.produto.produto_referencia
        preco = float(coleta.preco_atual)

        if referencia.id not in grupos:
            grupos[referencia.id] = {
                "referencia": referencia,
                "marca": referencia.marca.nome if referencia.marca else "-",
                "fabricante": (
                    referencia.fabricante_importador.nome
                    if referencia.fabricante_importador
                    else "-"
                ),
                "categoria": referencia.categoria.nome if referencia.categoria else "-",
                "sites": set(),
                "qtd_coletas": 0,
                "soma_precos": 0,
                "menor_preco": preco,
                "maior_preco": preco,
                "site_menor_preco": coleta.produto.site.nome,
                "site_maior_preco": coleta.produto.site.nome,
                "ultima_coleta": coleta.data_coleta,
            }

        grupo = grupos[referencia.id]

        grupo["sites"].add(coleta.produto.site.nome)
        grupo["qtd_coletas"] += 1
        grupo["soma_precos"] += preco

        if preco < grupo["menor_preco"]:
            grupo["menor_preco"] = preco
            grupo["site_menor_preco"] = coleta.produto.site.nome

        if preco > grupo["maior_preco"]:
            grupo["maior_preco"] = preco
            grupo["site_maior_preco"] = coleta.produto.site.nome

        if coleta.data_coleta > grupo["ultima_coleta"]:
            grupo["ultima_coleta"] = coleta.data_coleta

    resultado = []

    for grupo in grupos.values():
        preco_medio = grupo["soma_precos"] / grupo["qtd_coletas"]

        variacao_percentual = 0
        if grupo["menor_preco"] > 0:
            variacao_percentual = (
                (grupo["maior_preco"] - grupo["menor_preco"])
                / grupo["menor_preco"]
            ) * 100

        grupo["preco_medio"] = preco_medio
        grupo["variacao_percentual"] = variacao_percentual
        grupo["sites_texto"] = ", ".join(sorted(grupo["sites"]))

        resultado.append(grupo)

    resultado.sort(key=lambda item: item["preco_medio"], reverse=True)

    contexto = {
    "resultados": resultado,
    "total_resultados": len(resultado),
    "sites": SiteMonitorado.objects.all(),
    "marcas": Marca.objects.all(),
    "categorias": Categoria.objects.all(),
    "busca": busca,
    "site_id": site_id,
    "marca_id": marca_id,
    "categoria_id": categoria_id,
    "data_inicio": data_inicio,
    "data_fim": data_fim,
    "query_string": request.GET.urlencode(),
    }

    return render(request, "mercado/precos_referencias.html", contexto)

def exportar_precos_excel(request):
    coletas = ColetaProduto.objects.select_related(
        "produto",
        "produto__site",
        "produto__marca",
        "produto__categoria",
        "produto__produto_referencia",
        "produto__produto_referencia__marca",
        "produto__produto_referencia__categoria",
        "produto__produto_referencia__fabricante_importador",
    ).filter(
        produto__produto_referencia__isnull=False,
        preco_atual__isnull=False,
    )

    busca = request.GET.get("busca", "")
    site_id = request.GET.get("site", "")
    marca_id = request.GET.get("marca", "")
    categoria_id = request.GET.get("categoria", "")
    data_inicio = request.GET.get("data_inicio", "")
    data_fim = request.GET.get("data_fim", "")

    if busca:
        coletas = coletas.filter(
            Q(produto__produto_referencia__nome_referencia__icontains=busca)
            | Q(produto__produto_referencia__codigo_fabricante__icontains=busca)
            | Q(produto__produto_referencia__ean__icontains=busca)
            | Q(produto__nome_original__icontains=busca)
            | Q(produto__codigo_fabricante__icontains=busca)
            | Q(produto__ean__icontains=busca)
        )

    if site_id:
        coletas = coletas.filter(produto__site_id=site_id)

    if marca_id:
        coletas = coletas.filter(produto__produto_referencia__marca_id=marca_id)

    if categoria_id:
        coletas = coletas.filter(produto__produto_referencia__categoria_id=categoria_id)

    if data_inicio:
        coletas = coletas.filter(data_coleta__date__gte=data_inicio)

    if data_fim:
        coletas = coletas.filter(data_coleta__date__lte=data_fim)

    grupos = {}

    for coleta in coletas.order_by("produto__produto_referencia__nome_referencia", "data_coleta"):
        referencia = coleta.produto.produto_referencia
        preco = float(coleta.preco_atual)

        if referencia.id not in grupos:
            grupos[referencia.id] = {
                "referencia": referencia,
                "marca": referencia.marca.nome if referencia.marca else "-",
                "fabricante": (
                    referencia.fabricante_importador.nome
                    if referencia.fabricante_importador
                    else "-"
                ),
                "categoria": referencia.categoria.nome if referencia.categoria else "-",
                "sites": set(),
                "qtd_coletas": 0,
                "soma_precos": 0,
                "menor_preco": preco,
                "maior_preco": preco,
                "site_menor_preco": coleta.produto.site.nome,
                "site_maior_preco": coleta.produto.site.nome,
                "ultima_coleta": coleta.data_coleta,
            }

        grupo = grupos[referencia.id]

        grupo["sites"].add(coleta.produto.site.nome)
        grupo["qtd_coletas"] += 1
        grupo["soma_precos"] += preco

        if preco < grupo["menor_preco"]:
            grupo["menor_preco"] = preco
            grupo["site_menor_preco"] = coleta.produto.site.nome

        if preco > grupo["maior_preco"]:
            grupo["maior_preco"] = preco
            grupo["site_maior_preco"] = coleta.produto.site.nome

        if coleta.data_coleta > grupo["ultima_coleta"]:
            grupo["ultima_coleta"] = coleta.data_coleta

    resultado = []

    for grupo in grupos.values():
        preco_medio = grupo["soma_precos"] / grupo["qtd_coletas"]

        variacao_percentual = 0
        if grupo["menor_preco"] > 0:
            variacao_percentual = (
                (grupo["maior_preco"] - grupo["menor_preco"])
                / grupo["menor_preco"]
            ) * 100

        grupo["preco_medio"] = preco_medio
        grupo["variacao_percentual"] = variacao_percentual
        grupo["sites_texto"] = ", ".join(sorted(grupo["sites"]))

        resultado.append(grupo)

    resultado.sort(key=lambda item: item["preco_medio"], reverse=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Preco Medio"

    cabecalhos = [
        "Produto referência",
        "Marca",
        "Fabricante / Importador",
        "Categoria",
        "Sites",
        "Quantidade de coletas",
        "Menor preço",
        "Site menor preço",
        "Preço médio",
        "Maior preço",
        "Site maior preço",
        "Variação percentual",
        "Última coleta",
    ]

    sheet.append(cabecalhos)

    for item in resultado:
        sheet.append([
            item["referencia"].nome_referencia,
            item["marca"],
            item["fabricante"],
            item["categoria"],
            item["sites_texto"],
            item["qtd_coletas"],
            round(item["menor_preco"], 2),
            item["site_menor_preco"],
            round(item["preco_medio"], 2),
            round(item["maior_preco"], 2),
            item["site_maior_preco"],
            round(item["variacao_percentual"], 2),
            item["ultima_coleta"].strftime("%d/%m/%Y %H:%M"),
        ])

    for coluna in sheet.columns:
        largura = 12
        letra_coluna = coluna[0].column_letter

        for celula in coluna:
            if celula.value:
                largura = max(largura, len(str(celula.value)) + 2)

        sheet.column_dimensions[letra_coluna].width = min(largura, 60)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="preco_medio_produtos_referencia.xlsx"'

    workbook.save(response)
    return response

def exportar_ranking_excel(request):
    coletas = ColetaProduto.objects.select_related(
        "produto",
        "produto__site",
        "produto__marca",
        "produto__categoria",
        "produto__produto_referencia",
    ).order_by("-data_coleta")

    busca = request.GET.get("busca", "")
    site_id = request.GET.get("site", "")
    marca_id = request.GET.get("marca", "")
    categoria_id = request.GET.get("categoria", "")
    somente_disponiveis = request.GET.get("somente_disponiveis", "")

    if busca:
        coletas = coletas.filter(
            Q(produto__nome_original__icontains=busca)
            | Q(produto__codigo_fabricante__icontains=busca)
            | Q(produto__ean__icontains=busca)
            | Q(produto__produto_referencia__nome_referencia__icontains=busca)
            | Q(produto__produto_referencia__codigo_fabricante__icontains=busca)
            | Q(produto__produto_referencia__ean__icontains=busca)
        )

    if site_id:
        coletas = coletas.filter(produto__site_id=site_id)

    if marca_id:
        coletas = coletas.filter(produto__marca_id=marca_id)

    if categoria_id:
        coletas = coletas.filter(produto__categoria_id=categoria_id)

    if somente_disponiveis == "1":
        coletas = coletas.filter(disponivel=True)

    # Pega somente a última coleta de cada produto.
    ultimas_por_produto = {}
    for coleta in coletas.distinct():
        if coleta.produto_id not in ultimas_por_produto:
            ultimas_por_produto[coleta.produto_id] = coleta

    ranking = list(ultimas_por_produto.values())

    ranking.sort(
        key=lambda coleta: (
            coleta.ranking_geral if coleta.ranking_geral is not None else 999999,
            coleta.ranking_categoria if coleta.ranking_categoria is not None else 999999,
            coleta.preco_atual if coleta.preco_atual is not None else 999999999,
        )
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Ranking Mercado"

    cabecalhos = [
        "Ranking geral",
        "Ranking categoria",
        "Produto coletado",
        "Produto referência",
        "Site",
        "Marca",
        "Categoria",
        "Código site",
        "Código fabricante",
        "EAN",
        "Preço atual",
        "Preço antigo",
        "Desconto percentual",
        "Nota média",
        "Quantidade avaliações",
        "Disponível",
        "Status vínculo",
        "Data coleta",
        "URL",
    ]

    sheet.append(cabecalhos)

    for coleta in ranking:
        produto = coleta.produto

        sheet.append([
            coleta.ranking_geral,
            coleta.ranking_categoria,
            produto.nome_original,
            produto.produto_referencia.nome_referencia if produto.produto_referencia else "",
            produto.site.nome if produto.site else "",
            produto.marca.nome if produto.marca else "",
            produto.categoria.nome if produto.categoria else "",
            produto.codigo_site or "",
            produto.codigo_fabricante or "",
            produto.ean or "",
            float(coleta.preco_atual) if coleta.preco_atual is not None else "",
            float(coleta.preco_antigo) if coleta.preco_antigo is not None else "",
            float(coleta.desconto_percentual) if coleta.desconto_percentual is not None else "",
            float(coleta.nota_media) if coleta.nota_media is not None else "",
            coleta.quantidade_avaliacoes if coleta.quantidade_avaliacoes is not None else "",
            "Sim" if coleta.disponivel else "Não",
            produto.get_status_vinculo_display(),
            coleta.data_coleta.strftime("%d/%m/%Y %H:%M"),
            produto.url,
        ])

    for coluna in sheet.columns:
        largura = 12
        letra_coluna = coluna[0].column_letter

        for celula in coluna:
            if celula.value:
                largura = max(largura, len(str(celula.value)) + 2)

        sheet.column_dimensions[letra_coluna].width = min(largura, 60)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="ranking_mercado.xlsx"'

    workbook.save(response)
    return response

def produtos_pendentes_validacao(request):
    produtos = ProdutoColetado.objects.select_related(
        "site",
        "marca",
        "categoria",
        "produto_referencia",
    ).filter(
        Q(produto_referencia__isnull=True)
        | Q(status_vinculo="PENDENTE")
        | Q(status_vinculo="DIVERGENTE")
        | Q(status_vinculo="SEM_REFERENCIA")
    )

    busca = request.GET.get("busca", "")
    site_id = request.GET.get("site", "")
    marca_id = request.GET.get("marca", "")
    categoria_id = request.GET.get("categoria", "")
    status_vinculo = request.GET.get("status_vinculo", "")

    if busca:
        produtos = produtos.filter(
            Q(nome_original__icontains=busca)
            | Q(codigo_site__icontains=busca)
            | Q(codigo_fabricante__icontains=busca)
            | Q(ean__icontains=busca)
            | Q(url__icontains=busca)
        )

    if site_id:
        produtos = produtos.filter(site_id=site_id)

    if marca_id:
        produtos = produtos.filter(marca_id=marca_id)

    if categoria_id:
        produtos = produtos.filter(categoria_id=categoria_id)

    if status_vinculo:
        produtos = produtos.filter(status_vinculo=status_vinculo)

    produtos = produtos.order_by("site__nome", "marca__nome", "nome_original")

    contexto = {
        "produtos": produtos.distinct(),
        "total_produtos": produtos.distinct().count(),
        "sites": SiteMonitorado.objects.all(),
        "marcas": Marca.objects.all(),
        "categorias": Categoria.objects.all(),
        "status_choices": ProdutoColetado.STATUS_VINCULO_CHOICES,
        "busca": busca,
        "site_id": site_id,
        "marca_id": marca_id,
        "categoria_id": categoria_id,
        "status_vinculo": status_vinculo,
    }

    return render(request, "mercado/produtos_pendentes_validacao.html", contexto)

def importar_referencias_excel(request):
    if request.method == "POST":
        arquivo = request.FILES.get("arquivo_excel")

        if not arquivo:
            messages.error(request, "Nenhum arquivo foi enviado.")
            return redirect("mercado:importar_referencias_excel")

        if not arquivo.name.endswith(".xlsx"):
            messages.error(request, "Envie um arquivo Excel no formato .xlsx.")
            return redirect("mercado:importar_referencias_excel")

        try:
            workbook = load_workbook(arquivo, data_only=True)
            sheet = workbook.active

            cabecalhos = []
            for celula in sheet[1]:
                cabecalhos.append(str(celula.value).strip() if celula.value else "")

            mapa_colunas = {
                nome_coluna: indice
                for indice, nome_coluna in enumerate(cabecalhos)
            }

            colunas_obrigatorias = [
                "nome_referencia",
                "marca",
                "categoria",
            ]

            for coluna in colunas_obrigatorias:
                if coluna not in mapa_colunas:
                    messages.error(
                        request,
                        f"Coluna obrigatória ausente no Excel: {coluna}"
                    )
                    return redirect("mercado:importar_referencias_excel")

            total_linhas = 0
            total_criados = 0
            total_atualizados = 0
            total_erros = 0

            for numero_linha, linha in enumerate(sheet.iter_rows(min_row=2), start=2):
                def valor(nome_coluna):
                    indice = mapa_colunas.get(nome_coluna)

                    if indice is None:
                        return ""

                    celula = linha[indice].value

                    if celula is None:
                        return ""

                    return str(celula).strip()

                nome_referencia = valor("nome_referencia")
                marca_nome = valor("marca")
                categoria_nome = valor("categoria")

                if not nome_referencia or not marca_nome or not categoria_nome:
                    total_erros += 1
                    continue

                total_linhas += 1

                fabricante_nome = valor("fabricante_importador")
                tipo_fabricante = valor("tipo_fabricante") or "FABRICANTE"
                pais_origem = valor("pais_origem")

                codigo_fabricante = valor("codigo_fabricante")
                ean = valor("ean")
                url_oficial = valor("url_oficial")
                fonte_validacao = valor("fonte_validacao") or "SITE_FABRICANTE"
                status_validacao = valor("status_validacao") or "PENDENTE"
                observacao_validacao = valor("observacao_validacao")
                ativo_texto = valor("ativo").lower()

                ativo = True
                if ativo_texto in ["não", "nao", "false", "0", "inativo"]:
                    ativo = False

                marca, _ = Marca.objects.get_or_create(
                    nome=marca_nome.upper()
                )

                categoria, _ = Categoria.objects.get_or_create(
                    nome=categoria_nome
                )

                fabricante_importador = None
                if fabricante_nome:
                    fabricante_importador, _ = FabricanteImportador.objects.get_or_create(
                        nome=fabricante_nome,
                        defaults={
                            "tipo": tipo_fabricante,
                            "pais_origem": pais_origem,
                        },
                    )

                # Critério de identificação:
                # 1. EAN, se existir
                # 2. Código fabricante + marca
                # 3. Nome referência + marca
                referencia = None

                if ean:
                    referencia = ProdutoReferencia.objects.filter(ean=ean).first()

                if not referencia and codigo_fabricante:
                    referencia = ProdutoReferencia.objects.filter(
                        marca=marca,
                        codigo_fabricante=codigo_fabricante,
                    ).first()

                if not referencia:
                    referencia = ProdutoReferencia.objects.filter(
                        marca=marca,
                        nome_referencia=nome_referencia,
                    ).first()

                if referencia:
                    referencia.nome_referencia = nome_referencia
                    referencia.marca = marca
                    referencia.categoria = categoria
                    referencia.fabricante_importador = fabricante_importador
                    referencia.codigo_fabricante = codigo_fabricante
                    referencia.ean = ean
                    referencia.url_oficial = url_oficial
                    referencia.fonte_validacao = fonte_validacao
                    referencia.status_validacao = status_validacao
                    referencia.observacao_validacao = observacao_validacao
                    referencia.ativo = ativo
                    referencia.save()

                    total_atualizados += 1
                else:
                    ProdutoReferencia.objects.create(
                        nome_referencia=nome_referencia,
                        marca=marca,
                        categoria=categoria,
                        fabricante_importador=fabricante_importador,
                        codigo_fabricante=codigo_fabricante,
                        ean=ean,
                        url_oficial=url_oficial,
                        fonte_validacao=fonte_validacao,
                        status_validacao=status_validacao,
                        observacao_validacao=observacao_validacao,
                        ativo=ativo,
                    )

                    total_criados += 1

            messages.success(
                request,
                (
                    f"Importação concluída. "
                    f"Linhas processadas: {total_linhas}. "
                    f"Criados: {total_criados}. "
                    f"Atualizados: {total_atualizados}. "
                    f"Linhas ignoradas/erro: {total_erros}."
                )
            )

            return redirect("mercado:importar_referencias_excel")

        except Exception as erro:
            messages.error(request, f"Erro ao importar arquivo: {erro}")
            return redirect("mercado:importar_referencias_excel")

    return render(request, "mercado/importar_referencias_excel.html")

def normalizar_para_comparacao(texto):
    if not texto:
        return ""

    texto = str(texto).upper().strip()

    substituicoes = {
        ".": " ",
        ",": " ",
        "-": " ",
        "/": " ",
        "_": " ",
        "  ": " ",
    }

    for antigo, novo in substituicoes.items():
        texto = texto.replace(antigo, novo)

    return " ".join(texto.split())


def similaridade_texto(texto_a, texto_b):
    texto_a = normalizar_para_comparacao(texto_a)
    texto_b = normalizar_para_comparacao(texto_b)

    if not texto_a or not texto_b:
        return 0

    return SequenceMatcher(None, texto_a, texto_b).ratio()


def gerar_sugestao_para_produto(produto, referencias):
    melhor_sugestao = None

    produto_ean = normalizar_para_comparacao(produto.ean)
    produto_codigo = normalizar_para_comparacao(produto.codigo_fabricante)
    produto_marca = produto.marca.nome.upper() if produto.marca else ""

    for referencia in referencias:
        referencia_ean = normalizar_para_comparacao(referencia.ean)
        referencia_codigo = normalizar_para_comparacao(referencia.codigo_fabricante)
        referencia_marca = referencia.marca.nome.upper() if referencia.marca else ""

        criterio = None
        confianca = 0

        # 1. Melhor critério: EAN igual.
        if produto_ean and referencia_ean and produto_ean == referencia_ean:
            criterio = "EAN igual"
            confianca = 100

        # 2. Código fabricante igual + marca igual.
        elif (
            produto_codigo
            and referencia_codigo
            and produto_codigo == referencia_codigo
            and produto_marca
            and referencia_marca
            and produto_marca == referencia_marca
        ):
            criterio = "Código fabricante igual + marca igual"
            confianca = 95

        # 3. Código fabricante igual.
        elif produto_codigo and referencia_codigo and produto_codigo == referencia_codigo:
            criterio = "Código fabricante igual"
            confianca = 85

        # 4. Nome parecido + marca igual.
        else:
            similaridade = similaridade_texto(
                produto.nome_original,
                referencia.nome_referencia,
            )

            if (
                similaridade >= 0.70
                and produto_marca
                and referencia_marca
                and produto_marca == referencia_marca
            ):
                criterio = "Nome parecido + marca igual"
                confianca = round(similaridade * 100, 2)

            elif similaridade >= 0.82:
                criterio = "Nome muito parecido"
                confianca = round(similaridade * 100, 2)

        if criterio and confianca > 0:
            if not melhor_sugestao or confianca > melhor_sugestao["confianca"]:
                melhor_sugestao = {
                    "produto": produto,
                    "referencia": referencia,
                    "criterio": criterio,
                    "confianca": confianca,
                }

    return melhor_sugestao


def sugestoes_vinculo(request):
    produtos = ProdutoColetado.objects.select_related(
        "site",
        "marca",
        "categoria",
        "produto_referencia",
    ).filter(
        Q(produto_referencia__isnull=True)
        | Q(status_vinculo="PENDENTE")
        | Q(status_vinculo="DIVERGENTE")
        | Q(status_vinculo="SEM_REFERENCIA")
    )

    referencias = ProdutoReferencia.objects.select_related(
        "marca",
        "categoria",
        "fabricante_importador",
    ).filter(
        ativo=True
    )

    busca = request.GET.get("busca", "")
    minimo_confianca = request.GET.get("minimo_confianca", "70")

    if busca:
        produtos = produtos.filter(
            Q(nome_original__icontains=busca)
            | Q(codigo_fabricante__icontains=busca)
            | Q(ean__icontains=busca)
            | Q(url__icontains=busca)
        )

    try:
        minimo_confianca_num = float(minimo_confianca)
    except Exception:
        minimo_confianca_num = 70

    sugestoes = []

    for produto in produtos.order_by("site__nome", "marca__nome", "nome_original"):
        sugestao = gerar_sugestao_para_produto(produto, referencias)

        if sugestao and sugestao["confianca"] >= minimo_confianca_num:
            sugestoes.append(sugestao)

    sugestoes.sort(key=lambda item: item["confianca"], reverse=True)

    contexto = {
        "sugestoes": sugestoes,
        "total_sugestoes": len(sugestoes),
        "busca": busca,
        "minimo_confianca": minimo_confianca,
    }

    return render(request, "mercado/sugestoes_vinculo.html", contexto)


def aplicar_sugestao_vinculo(request, produto_id, referencia_id):
    if request.method != "POST":
        return redirect("mercado:sugestoes_vinculo")

    produto = get_object_or_404(ProdutoColetado, id=produto_id)
    referencia = get_object_or_404(ProdutoReferencia, id=referencia_id)

    produto.produto_referencia = referencia
    produto.status_vinculo = "VINCULADO"

    # Se o produto coletado estiver sem marca/categoria, aproveita a referência.
    if not produto.marca and referencia.marca:
        produto.marca = referencia.marca

    if not produto.categoria and referencia.categoria:
        produto.categoria = referencia.categoria

    # Se estiver sem código/EAN, aproveita a referência oficial.
    if not produto.codigo_fabricante and referencia.codigo_fabricante:
        produto.codigo_fabricante = referencia.codigo_fabricante

    if not produto.ean and referencia.ean:
        produto.ean = referencia.ean

    produto.save()

    messages.success(
        request,
        f"Produto '{produto.nome_original}' vinculado à referência '{referencia.nome_referencia}'."
    )

    return redirect("mercado:sugestoes_vinculo")

def aplicar_sugestoes_alta_confianca(request):
    if request.method != "POST":
        return redirect("mercado:sugestoes_vinculo")

    try:
        minimo_confianca = float(request.POST.get("minimo_confianca_lote", "95"))
    except Exception:
        minimo_confianca = 95

    produtos = ProdutoColetado.objects.select_related(
        "site",
        "marca",
        "categoria",
        "produto_referencia",
    ).filter(
        Q(produto_referencia__isnull=True)
        | Q(status_vinculo="PENDENTE")
        | Q(status_vinculo="DIVERGENTE")
        | Q(status_vinculo="SEM_REFERENCIA")
    )

    referencias = ProdutoReferencia.objects.select_related(
        "marca",
        "categoria",
        "fabricante_importador",
    ).filter(
        ativo=True
    )

    criterios_permitidos = {
        "EAN igual",
        "Código fabricante igual + marca igual",
    }

    total_vinculados = 0
    total_ignorados = 0

    for produto in produtos:
        sugestao = gerar_sugestao_para_produto(produto, referencias)

        if not sugestao:
            total_ignorados += 1
            continue

        if sugestao["confianca"] < minimo_confianca:
            total_ignorados += 1
            continue

        if sugestao["criterio"] not in criterios_permitidos:
            total_ignorados += 1
            continue

        referencia = sugestao["referencia"]

        produto.produto_referencia = referencia
        produto.status_vinculo = "VINCULADO"

        if not produto.marca and referencia.marca:
            produto.marca = referencia.marca

        if not produto.categoria and referencia.categoria:
            produto.categoria = referencia.categoria

        if not produto.codigo_fabricante and referencia.codigo_fabricante:
            produto.codigo_fabricante = referencia.codigo_fabricante

        if not produto.ean and referencia.ean:
            produto.ean = referencia.ean

        produto.save()
        total_vinculados += 1

    messages.success(
        request,
        (
            f"Aplicação em lote concluída. "
            f"Produtos vinculados: {total_vinculados}. "
            f"Produtos ignorados: {total_ignorados}."
        )
    )

    return redirect("mercado:sugestoes_vinculo")

def forca_marcas(request):
    produtos = ProdutoColetado.objects.select_related(
        "site",
        "marca",
        "categoria",
    ).filter(
        ativo=True,
        marca__isnull=False,
    )

    busca = request.GET.get("busca", "")
    site_id = request.GET.get("site", "")
    categoria_id = request.GET.get("categoria", "")

    if busca:
        produtos = produtos.filter(
            Q(nome_original__icontains=busca)
            | Q(marca__nome__icontains=busca)
            | Q(codigo_site__icontains=busca)
            | Q(codigo_fabricante__icontains=busca)
        )

    if site_id:
        produtos = produtos.filter(site_id=site_id)

    if categoria_id:
        produtos = produtos.filter(categoria_id=categoria_id)

    marcas = {}

    for produto in produtos:
        if not produto.marca:
            continue

        ultima_coleta = produto.coletas.order_by("-data_coleta").first()

        if not ultima_coleta:
            continue

        marca_nome = produto.marca.nome

        if marca_nome not in marcas:
            marcas[marca_nome] = {
                "marca": produto.marca,
                "qtd_produtos": 0,
                "qtd_vinculados": 0,
                "qtd_disponiveis": 0,
                "soma_preco_atual": Decimal("0"),
                "qtd_preco_atual": 0,
                "soma_preco_prazo": Decimal("0"),
                "qtd_preco_prazo": 0,
                "soma_nota": Decimal("0"),
                "qtd_nota": 0,
                "total_avaliacoes": 0,
                "soma_ranking": Decimal("0"),
                "qtd_ranking": 0,
            }

        item = marcas[marca_nome]

        item["qtd_produtos"] += 1

        if produto.produto_referencia_id:
            item["qtd_vinculados"] += 1

        if ultima_coleta.disponivel:
            item["qtd_disponiveis"] += 1

        if ultima_coleta.preco_atual:
            item["soma_preco_atual"] += ultima_coleta.preco_atual
            item["qtd_preco_atual"] += 1

        if ultima_coleta.preco_prazo:
            item["soma_preco_prazo"] += ultima_coleta.preco_prazo
            item["qtd_preco_prazo"] += 1

        if ultima_coleta.nota_media:
            item["soma_nota"] += ultima_coleta.nota_media
            item["qtd_nota"] += 1

        if ultima_coleta.quantidade_avaliacoes:
            item["total_avaliacoes"] += ultima_coleta.quantidade_avaliacoes

        if ultima_coleta.ranking_geral:
            item["soma_ranking"] += Decimal(ultima_coleta.ranking_geral)
            item["qtd_ranking"] += 1

    ranking_marcas = []

    for dados in marcas.values():
        preco_medio_atual = None
        preco_medio_prazo = None
        nota_media = None
        ranking_medio = None

        if dados["qtd_preco_atual"]:
            preco_medio_atual = dados["soma_preco_atual"] / dados["qtd_preco_atual"]

        if dados["qtd_preco_prazo"]:
            preco_medio_prazo = dados["soma_preco_prazo"] / dados["qtd_preco_prazo"]

        if dados["qtd_nota"]:
            nota_media = dados["soma_nota"] / dados["qtd_nota"]

        if dados["qtd_ranking"]:
            ranking_medio = dados["soma_ranking"] / dados["qtd_ranking"]

        ranking_marcas.append({
            "marca": dados["marca"],
            "qtd_produtos": dados["qtd_produtos"],
            "qtd_vinculados": dados["qtd_vinculados"],
            "qtd_disponiveis": dados["qtd_disponiveis"],
            "preco_medio_atual": preco_medio_atual,
            "preco_medio_prazo": preco_medio_prazo,
            "nota_media": nota_media,
            "total_avaliacoes": dados["total_avaliacoes"],
            "ranking_medio": ranking_medio,
        })

    ranking_marcas.sort(
        key=lambda item: (
            -(item["qtd_produtos"] or 0),
            item["ranking_medio"] if item["ranking_medio"] is not None else Decimal("999999"),
        )
    )

    contexto = {
        "ranking_marcas": ranking_marcas,
        "total_marcas": len(ranking_marcas),
        "sites": SiteMonitorado.objects.all(),
        "categorias": Categoria.objects.all(),
        "busca": busca,
        "site_id": site_id,
        "categoria_id": categoria_id,
    }

    return render(request, "mercado/forca_marcas.html", contexto)